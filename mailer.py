# -*- coding: utf-8 -*-
"""
분석 결과를 뉴스레터 형식 HTML 이메일로 발송
.env 필요: MAIL_FROM, MAIL_TO, SMTP_HOST, SMTP_PORT, SMTP_USER, SMTP_PASS
"""
import os
import smtplib
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from dotenv import load_dotenv

load_dotenv()


# ── HTML 생성 ─────────────────────────────────────────────────

def build_html(results: list[dict], date_str: str, total_collected: int,
               screened_all: list[dict] = None,
               company_name: str = "", keywords: list[str] = None) -> str:
    high = [r for r in results if r.get("analysis", {}).get("verdict") == "HIGH"]
    med  = [r for r in results if r.get("analysis", {}).get("verdict") == "MED"]

    date_label = f"{date_str[:4]}년 {date_str[4:6]}월 {date_str[6:]}일"
    kw_str = ", ".join(keywords) if keywords else ""

    greeting = _greeting(company_name, date_label, kw_str, total_collected)

    # HIGH는 강조 카드, MED는 일반 카드
    high_cards = "".join(_high_card(r) for r in high)
    med_section = _section("🟡 검토 고려", "#d68910", "#fffbf0", med) if med else ""

    if not high and not med:
        body = """<div style="text-align:center;padding:48px 0;color:#888;font-size:15px;">
          오늘은 해당하는 공고가 없습니다.</div>"""
    else:
        high_section = f"""
        <div style="margin-top:8px;">
          <div style="font-size:16px;font-weight:700;color:#c0392b;margin-bottom:16px;
               padding-bottom:8px;border-bottom:2px solid #c0392b;">🔴 즉시 검토</div>
          {high_cards}
        </div>""" if high else ""
        body = high_section + med_section

    table_section = _all_table(screened_all) if screened_all else ""

    return f"""<!DOCTYPE html>
<html lang="ko">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>나라장터 AI 공고 모니터링</title>
<style>
@media screen and (max-width:620px){{
  .wrap{{width:100%!important;}}
  .pad{{padding-left:16px!important;padding-right:16px!important;}}
  .hpad{{padding:20px 16px 16px!important;}}
  .mob-block{{display:block!important;width:100%!important;}}
  .mob-badge{{width:25%!important;}}
  .score-num{{font-size:24px!important;}}
  .title-txt{{font-size:14px!important;}}
  .meta-txt{{font-size:12px!important;}}
  .col2-left{{display:block!important;width:100%!important;padding-right:0!important;padding-bottom:10px!important;}}
  .col2-right{{display:block!important;width:100%!important;padding-left:0!important;border-left:none!important;}}
}}
</style>
</head>
<body style="margin:0;padding:0;background:#f0f2f5;font-family:'Apple SD Gothic Neo','Malgun Gothic',sans-serif;">

<table width="100%" cellpadding="0" cellspacing="0" style="background:#f0f2f5;padding:16px 0;">
<tr><td align="center">
<table class="wrap" width="680" cellpadding="0" cellspacing="0" style="max-width:680px;width:100%;">

  <!-- 헤더 -->
  <tr><td class="hpad" style="background:linear-gradient(135deg,#1a1f36 0%,#2d3561 100%);border-radius:12px 12px 0 0;padding:28px 36px 20px;">
    <div style="color:#7c8cf8;font-size:11px;letter-spacing:2px;text-transform:uppercase;margin-bottom:6px;">NARA MONITOR</div>
    <div style="color:#fff;font-size:20px;font-weight:700;margin-bottom:4px;">나라장터 AI 공고 분석 리포트</div>
    <div style="color:#a0aec0;font-size:12px;">{date_label} 기준</div>
  </td></tr>

  <!-- 인사말 -->
  {greeting}

  <!-- 요약 배지 -->
  <tr><td class="pad" style="background:#fff;padding:16px 36px 20px;border-bottom:1px solid #e8ecf0;">
    <table width="100%" cellpadding="0" cellspacing="0"><tr>
      {_badge("수집", total_collected, "#4a5568", "건")}
      {_badge("즉시검토", len(high), "#c0392b", "건")}
      {_badge("검토고려", len(med), "#d68910", "건")}
      {_badge("분석완료", len(results), "#2980b9", "건")}
    </tr></table>
  </td></tr>

  <!-- 본문 -->
  <tr><td class="pad" style="background:#fff;padding:16px 36px 32px;">
    {body}
  </td></tr>

  <!-- 전체 수집 목록 표 -->
  {table_section}

  <!-- 푸터 -->
  <tr><td style="background:#2d3561;border-radius:0 0 12px 12px;padding:18px 40px;text-align:center;">
    <div style="color:#7c8cf8;font-size:12px;">Nara Monitor Service · 자동 생성된 보고서입니다</div>
  </td></tr>

</table>
</td></tr>
</table>
</body>
</html>"""


def _greeting(company_name: str, date_label: str, kw_str: str, total: int) -> str:
    name_part = f"<b>{company_name}</b> 님" if company_name else "안녕하세요"
    kw_part   = f"<b>{kw_str}</b> 키워드로" if kw_str else ""
    return f"""
  <tr><td class="pad" style="background:#fff;padding:20px 36px 0;">
    <div style="font-size:14px;color:#2d3748;line-height:1.9;border-left:3px solid #7c8cf8;padding-left:14px;">
      안녕하세요, {name_part}.<br>
      <b>{date_label}</b> 나라장터에 등록된 공고 중 {kw_part} 검색·분석한 결과를 전달드립니다.<br>
      오늘 총 <b>{total}건</b>이 수집되었으며, AI가 클라비 사업 역량과의 관련성을 검토했습니다.
    </div>
  </td></tr>"""


def _high_card(r: dict) -> str:
    """HIGH 전용 강조 카드 — 마감일 배지 부각, 점수 크게"""
    a        = r.get("analysis", {})
    score    = a.get("score", "?")
    summary  = a.get("summary", "")
    matches  = a.get("match_points", [])
    gaps     = a.get("gap_points", [])
    action   = a.get("action", "")
    source   = r.get("source", "")
    biz_type = r.get("type", "")
    title    = r.get("title", "")
    agency   = r.get("agency", "")
    amount   = r.get("amount", "")
    close    = r.get("close_date", "")
    url      = r.get("detail_url", "")

    try:
        amt_str = f"{int(amount):,}원" if amount and str(amount).isdigit() and int(amount) > 0 else "금액 미공개"
    except Exception:
        amt_str = "금액 미공개"

    close_str = _fmt_date(close) if close else "마감일 미정"

    match_html = "".join(f'<li style="margin-bottom:5px;">{m}</li>' for m in matches) \
                 if matches else '<li style="color:#999;">-</li>'
    gap_html   = "".join(f'<li style="margin-bottom:5px;">{g}</li>' for g in gaps) \
                 if gaps else '<li style="color:#999;">-</li>'

    detail_btn = f"""<a href="{url}" style="display:inline-block;padding:8px 20px;background:#c0392b;color:#fff;border-radius:6px;font-size:12px;text-decoration:none;font-weight:700;">공고 바로가기 →</a>""" if url else ""

    return f"""
    <div style="background:#fff;border:1.5px solid #f5c6c6;border-radius:10px;margin-bottom:20px;overflow:hidden;box-shadow:0 2px 8px rgba(192,57,43,0.08);">

      <!-- 상단 컬러 바 + 점수 -->
      <table width="100%" cellpadding="0" cellspacing="0">
      <tr>
        <td style="background:linear-gradient(90deg,#c0392b 0%,#e74c3c 100%);padding:14px 20px;vertical-align:middle;">
          <span style="background:rgba(255,255,255,0.2);color:#fff;font-size:11px;font-weight:700;padding:3px 10px;border-radius:4px;">{source} · {biz_type}</span>
        </td>
        <td style="background:linear-gradient(90deg,#c0392b 0%,#e74c3c 100%);padding:14px 20px;text-align:right;vertical-align:middle;white-space:nowrap;width:90px;">
          <span style="color:#fff;font-size:32px;font-weight:800;line-height:1;">{score}</span>
          <span style="color:rgba(255,255,255,0.7);font-size:11px;"> / 100</span>
        </td>
      </tr>
      </table>

      <!-- 제목 + 메타 -->
      <div style="padding:16px 20px 0;">
        <div style="font-size:16px;font-weight:700;color:#1a202c;line-height:1.5;margin-bottom:12px;">{title}</div>

        <!-- 1줄: 기관 + 금액 (강조) -->
        <table width="100%" cellpadding="0" cellspacing="0" style="margin-bottom:8px;"><tr>
          <td style="font-size:15px;font-weight:700;color:#1a202c;">🏢 {agency}</td>
          <td style="font-size:16px;font-weight:800;color:#c0392b;text-align:right;white-space:nowrap;">💰 {amt_str}</td>
        </tr></table>

        <!-- 2줄: 마감일 배지 -->
        <div style="margin-bottom:12px;">
          <span style="display:inline-block;background:#c0392b;color:#fff;font-size:13px;font-weight:700;padding:5px 16px;border-radius:20px;">
            📅 마감 {close_str}
          </span>
        </div>
      </div>

      <!-- 구분선 -->
      <div style="border-top:1px solid #fde8e8;margin:0 20px;"></div>

      <!-- 요약 -->
      <div style="padding:12px 20px;font-size:13px;color:#2d3748;line-height:1.7;background:#fff9f9;">
        {summary}
      </div>

      <!-- 강점 / 리스크 -->
      <table width="100%" cellpadding="0" cellspacing="0" style="padding:0 20px 12px;">
      <tr>
        <td class="col2-left" width="50%" style="vertical-align:top;padding:12px 8px 0 0;">
          <div style="font-size:12px;font-weight:700;color:#27ae60;margin-bottom:6px;">✅ 강점</div>
          <ul style="margin:0;padding-left:16px;font-size:12px;color:#2d3748;line-height:1.7;">{match_html}</ul>
        </td>
        <td class="col2-right" width="50%" style="vertical-align:top;padding:12px 0 0 8px;border-left:1px dashed #fde8e8;">
          <div style="font-size:12px;font-weight:700;color:#e74c3c;margin-bottom:6px;">⚠️ 리스크</div>
          <ul style="margin:0;padding-left:16px;font-size:12px;color:#2d3748;line-height:1.7;">{gap_html}</ul>
        </td>
      </tr>
      </table>

      <!-- 권장 행동 + 버튼 -->
      <table width="100%" cellpadding="0" cellspacing="0"
             style="background:#fff3f3;border-top:1px solid #fde8e8;padding:0;">
      <tr>
        <td style="padding:12px 20px;font-size:12px;color:#2d3748;vertical-align:middle;">
          <span style="font-weight:700;color:#c0392b;">▶ 권장 행동&nbsp;</span>{action}
        </td>
        <td style="padding:12px 20px;text-align:right;vertical-align:middle;white-space:nowrap;">
          {detail_btn}
        </td>
      </tr>
      </table>

    </div>"""


def _badge(label: str, value: int, color: str, unit: str) -> str:
    return f"""
    <td style="text-align:center;padding:8px 0;">
      <div style="color:{color};font-size:26px;font-weight:700;">{value}<span style="font-size:14px;margin-left:2px;">{unit}</span></div>
      <div style="color:#888;font-size:12px;margin-top:2px;">{label}</div>
    </td>"""


def _section(title: str, accent: str, bg: str, items: list[dict]) -> str:
    cards = "".join(_card(r, accent, bg) for r in items)
    return f"""
    <div style="margin-top:28px;">
      <div style="font-size:16px;font-weight:700;color:{accent};margin-bottom:14px;padding-bottom:8px;border-bottom:2px solid {accent};">{title}</div>
      {cards}
    </div>"""


def _card(r: dict, accent: str, bg: str) -> str:
    a = r.get("analysis", {})
    score   = a.get("score", "?")
    summary = a.get("summary", "")
    matches = a.get("match_points", [])
    gaps    = a.get("gap_points", [])
    action  = a.get("action", "")

    source  = r.get("source", "")
    biz_type = r.get("type", "")
    title   = r.get("title", "")
    agency  = r.get("agency", "")
    amount  = r.get("amount", "")
    close   = r.get("close_date", "")
    url     = r.get("detail_url", "")
    screen_reason = r.get("screen_reason", "")

    try:
        amt_str = f"{int(amount):,}원" if amount and str(amount).isdigit() and int(amount) > 0 else "금액 미공개"
    except Exception:
        amt_str = "금액 미공개"

    close_str = _fmt_date(close) if close else "마감일 미정"

    match_html = "".join(
        f'<li style="margin-bottom:4px;color:#2d3748;">{m}</li>' for m in matches
    ) if matches else '<li style="color:#888;">-</li>'

    gap_html = "".join(
        f'<li style="margin-bottom:4px;color:#2d3748;">{g}</li>' for g in gaps
    ) if gaps else '<li style="color:#888;">-</li>'

    detail_btn = f"""<a href="{url}" style="display:inline-block;margin-top:12px;padding:7px 16px;background:{accent};color:#fff;border-radius:6px;font-size:12px;text-decoration:none;font-weight:600;">공고 바로가기 →</a>""" if url else ""

    return f"""
    <div style="background:{bg};border:1px solid #e2e8f0;border-left:4px solid {accent};border-radius:8px;padding:20px 22px;margin-bottom:16px;">

      <!-- 타이틀 행 -->
      <table width="100%" cellpadding="0" cellspacing="0"><tr>
        <td style="vertical-align:top;">
          <span style="display:inline-block;background:{accent};color:#fff;font-size:11px;font-weight:700;padding:2px 8px;border-radius:4px;margin-bottom:8px;">{source} · {biz_type}</span>
          <div style="font-size:15px;font-weight:700;color:#1a202c;line-height:1.5;">{title}</div>
        </td>
        <td style="vertical-align:top;text-align:right;white-space:nowrap;padding-left:12px;">
          <div style="font-size:28px;font-weight:800;color:{accent};">{score}</div>
          <div style="font-size:10px;color:#888;margin-top:-2px;">/ 100점</div>
        </td>
      </tr></table>

      <!-- 메타 -->
      <div style="margin-top:10px;font-size:12px;color:#4a5568;">
        🏢 {agency} &nbsp;|&nbsp; 💰 {amt_str} &nbsp;|&nbsp; 📅 마감 {close_str}
      </div>

      <!-- 요약 -->
      <div style="margin-top:12px;font-size:13px;color:#2d3748;line-height:1.7;padding:10px 14px;background:rgba(255,255,255,0.7);border-radius:6px;">
        {summary}
      </div>

      <!-- 강점 / 리스크 -->
      <table width="100%" cellpadding="0" cellspacing="0" style="margin-top:14px;">
      <tr>
        <td class="col2-left" width="50%" style="vertical-align:top;padding-right:8px;">
          <div style="font-size:12px;font-weight:700;color:#27ae60;margin-bottom:6px;">✅ 강점</div>
          <ul style="margin:0;padding-left:16px;font-size:12px;line-height:1.7;">{match_html}</ul>
        </td>
        <td class="col2-right" width="50%" style="vertical-align:top;padding-left:8px;">
          <div style="font-size:12px;font-weight:700;color:#e74c3c;margin-bottom:6px;">⚠️ 리스크</div>
          <ul style="margin:0;padding-left:16px;font-size:12px;line-height:1.7;">{gap_html}</ul>
        </td>
      </tr>
      </table>

      <!-- 행동 지침 -->
      <div style="margin-top:14px;padding:10px 14px;background:rgba(255,255,255,0.8);border-radius:6px;font-size:12px;color:#2d3748;">
        <span style="font-weight:700;color:{accent};">▶ 권장 행동</span>&nbsp; {action}
      </div>

      {detail_btn}
    </div>"""


def _all_table(screened: list[dict]) -> str:
    """전체 수집 목록을 컴팩트 표로 렌더링 (이메일 하단 첨부)"""
    if not screened:
        return ""

    REL_STYLE = {
        "HIGH": ("🔴", "#c0392b", "#fff5f5"),
        "MED":  ("🟡", "#d68910", "#fffbf0"),
        "LOW":  ("⚪", "#718096", "#f7f8fa"),
    }

    rows = ""
    for i, n in enumerate(screened):
        rel = n.get("relevance", "LOW")
        icon, color, bg = REL_STYLE.get(rel, REL_STYLE["LOW"])
        title = n.get("title", "")
        source = n.get("source", "")
        biz_type = n.get("type", "")
        agency = n.get("agency", "")
        amount = n.get("amount", "")
        close = n.get("close_date", "")
        url = n.get("detail_url", "")
        reason = n.get("screen_reason", "")

        try:
            amt_str = f"{int(amount):,}" if amount and str(amount).isdigit() and int(amount) > 0 else "-"
        except Exception:
            amt_str = "-"

        close_str = _fmt_date(close) if close else "-"
        title_cell = f'<a href="{url}" style="color:#2b6cb0;text-decoration:none;">{title}</a>' if url else title

        row_bg = "#ffffff" if i % 2 == 0 else "#f9fafb"
        rows += f"""
        <tr style="background:{row_bg};">
          <td style="padding:7px 10px;font-size:11px;color:{color};white-space:nowrap;">{icon} {rel}</td>
          <td style="padding:7px 6px;font-size:11px;color:#4a5568;white-space:nowrap;">{source}<br>{biz_type}</td>
          <td style="padding:7px 10px;font-size:12px;color:#1a202c;line-height:1.4;">{title_cell}</td>
          <td style="padding:7px 8px;font-size:11px;color:#4a5568;white-space:nowrap;">{agency}</td>
          <td style="padding:7px 8px;font-size:11px;color:#2d3748;text-align:right;white-space:nowrap;">{amt_str}</td>
          <td style="padding:7px 8px;font-size:11px;color:#4a5568;white-space:nowrap;">{close_str}</td>
          <td style="padding:7px 8px;font-size:11px;color:#718096;line-height:1.4;">{reason}</td>
        </tr>"""

    return f"""
  <tr><td style="background:#f0f2f5;padding:16px 0 0;"></td></tr>
  <tr><td style="background:#fff;border-radius:12px;padding:20px 40px 28px;">
    <details>
      <summary style="cursor:pointer;font-size:15px;font-weight:700;color:#2d3748;
                      padding:4px 0;list-style:none;user-select:none;outline:none;">
        <span style="display:inline-block;margin-right:6px;font-size:13px;">▶</span>
        📋 전체 수집 목록 ({len(screened)}건)
        <span style="font-size:12px;color:#888;font-weight:400;margin-left:8px;">— 클릭하여 펼치기</span>
      </summary>
      <div style="margin-top:14px;padding-top:14px;border-top:1px solid #e2e8f0;overflow-x:auto;">
      <table width="100%" cellpadding="0" cellspacing="0" style="border-collapse:collapse;font-size:12px;min-width:600px;">
        <thead>
          <tr style="background:#edf2f7;">
            <th style="padding:8px 10px;text-align:left;font-size:11px;color:#4a5568;white-space:nowrap;">관련성</th>
            <th style="padding:8px 6px;text-align:left;font-size:11px;color:#4a5568;white-space:nowrap;">출처</th>
            <th style="padding:8px 10px;text-align:left;font-size:11px;color:#4a5568;">공고명</th>
            <th style="padding:8px 8px;text-align:left;font-size:11px;color:#4a5568;">기관</th>
            <th style="padding:8px 8px;text-align:right;font-size:11px;color:#4a5568;">금액(원)</th>
            <th style="padding:8px 8px;text-align:left;font-size:11px;color:#4a5568;white-space:nowrap;">마감</th>
            <th style="padding:8px 8px;text-align:left;font-size:11px;color:#4a5568;">선별 이유</th>
          </tr>
        </thead>
        <tbody>{rows}</tbody>
      </table>
      </div>
    </details>
  </td></tr>"""


def _fmt_date(dt_str: str) -> str:
    """'20260530 1800' 또는 '2026-05-30' 등 → '05/30 18:00'"""
    if not dt_str:
        return "미정"
    dt_str = dt_str.strip()
    try:
        if len(dt_str) == 12 and dt_str[:8].isdigit():
            return f"{dt_str[4:6]}/{dt_str[6:8]} {dt_str[8:10]}:{dt_str[10:12]}"
        if len(dt_str) == 8 and dt_str.isdigit():
            return f"{dt_str[4:6]}/{dt_str[6:8]}"
    except Exception:
        pass
    return dt_str[:10]


# ── 이메일 발송 ───────────────────────────────────────────────

def send(html: str, date_str: str, subject: str = None, excel_path: str = None,
         mail_to: str = None) -> bool:
    from email.mime.base import MIMEBase
    from email import encoders

    host  = os.getenv("SMTP_HOST", "smtp.gmail.com")
    port  = int(os.getenv("SMTP_PORT", "587"))
    user  = os.getenv("SMTP_USER")
    pw    = os.getenv("SMTP_PASS")
    from_ = os.getenv("MAIL_FROM", user)
    to_   = mail_to or os.getenv("MAIL_TO")

    if not all([user, pw, to_]):
        print("  [메일 미설정] .env에 SMTP_USER, SMTP_PASS, MAIL_TO 필요")
        return False

    date_label = f"{date_str[:4]}.{date_str[4:6]}.{date_str[6:]}"
    subject = subject or f"[나라장터 모니터] {date_label} AI 공고 분석 리포트"

    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = from_
    msg["To"]      = to_

    msg.attach(MIMEText(html, "html", "utf-8"))

    if excel_path and os.path.exists(excel_path):
        with open(excel_path, "rb") as f:
            part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            part.set_payload(f.read())
        encoders.encode_base64(part)
        fname = os.path.basename(excel_path)
        part.add_header("Content-Disposition", "attachment", filename=fname)
        msg.attach(part)

    try:
        if port == 465:
            with smtplib.SMTP_SSL(host, port) as smtp:
                smtp.login(user, pw)
                smtp.sendmail(from_, [a.strip() for a in to_.split(",")], msg.as_string())
        else:
            with smtplib.SMTP(host, port) as smtp:
                smtp.ehlo()
                smtp.starttls()
                smtp.login(user, pw)
                smtp.sendmail(from_, [a.strip() for a in to_.split(",")], msg.as_string())
        print(f"  [발송 완료] → {to_}" + (f" (첨부: {os.path.basename(excel_path)})" if excel_path else ""))
        return True
    except Exception as e:
        print(f"  [발송 실패] {e}")
        return False


def save_html(html: str, date_str: str, path: str = None) -> str:
    """발송 전 로컬에 HTML 저장 (확인용)"""
    if path is None:
        path = f"report_{date_str}.html"
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    print(f"  [HTML 저장] {path}")
    return path


def save_excel(results: list[dict], screened_all: list[dict],
               date_str: str, path: str = None) -> str:
    """분석결과 + 전체수집 두 시트로 Excel 저장"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        print("  [Excel 생략] openpyxl 미설치 (pip install openpyxl)")
        return ""

    if path is None:
        path = f"report_{date_str}.xlsx"

    wb = openpyxl.Workbook()

    # ── 시트1: 분석결과 (HIGH/MED) ──────────────────────────────
    ws1 = wb.active
    ws1.title = "분석결과"

    hdr1 = ["점수", "판정", "출처", "구분", "공고명", "기관", "금액(원)", "마감일",
            "요약", "강점1", "강점2", "강점3", "리스크1", "리스크2", "권장행동", "공고URL"]
    _write_header(ws1, hdr1)

    for r in results:
        a = r.get("analysis", {})
        matches = a.get("match_points", [])
        gaps    = a.get("gap_points", [])
        amt = r.get("amount", "")
        try:
            amt = int(amt) if amt and str(amt).isdigit() and int(amt) > 0 else ""
        except Exception:
            amt = ""
        ws1.append([
            a.get("score", ""),
            a.get("verdict", ""),
            r.get("source", ""),
            r.get("type", ""),
            r.get("title", ""),
            r.get("agency", ""),
            amt,
            r.get("close_date", ""),
            a.get("summary", ""),
            matches[0] if len(matches) > 0 else "",
            matches[1] if len(matches) > 1 else "",
            matches[2] if len(matches) > 2 else "",
            gaps[0] if len(gaps) > 0 else "",
            gaps[1] if len(gaps) > 1 else "",
            a.get("action", ""),
            r.get("detail_url", ""),
        ])

    _style_sheet(ws1, col_widths=[6, 7, 8, 6, 40, 20, 14, 14, 50, 30, 30, 30, 30, 30, 40, 30])

    # ── 시트2: 전체수집 ───────────────────────────────────────
    ws2 = wb.create_sheet("전체수집")
    hdr2 = ["관련성", "출처", "구분", "공고명", "기관", "수요기관", "금액(원)",
            "등록일", "마감일", "계약방식", "선별이유", "공고번호", "첨부파일수", "공고URL"]
    _write_header(ws2, hdr2)

    REL_COLORS = {"HIGH": "FFC0C0", "MED": "FFF3C0", "LOW": "F0F0F0"}
    for r in screened_all:
        rel = r.get("relevance", "LOW")
        amt = r.get("amount", "")
        try:
            amt = int(amt) if amt and str(amt).isdigit() and int(amt) > 0 else ""
        except Exception:
            amt = ""
        row_idx = ws2.max_row + 1
        ws2.append([
            rel,
            r.get("source", ""),
            r.get("type", ""),
            r.get("title", ""),
            r.get("agency", ""),
            r.get("demand_agency", ""),
            amt,
            r.get("reg_date", ""),
            r.get("close_date", ""),
            r.get("contract_method", ""),
            r.get("screen_reason", ""),
            r.get("notice_no", ""),
            len(r.get("files", [])),
            r.get("detail_url", ""),
        ])
        fill_color = REL_COLORS.get(rel, "F0F0F0")
        ws2.cell(row_idx, 1).fill = PatternFill("solid", fgColor=fill_color)

    _style_sheet(ws2, col_widths=[7, 8, 6, 42, 20, 20, 14, 14, 14, 12, 40, 16, 8, 30])

    wb.save(path)
    print(f"  [Excel 저장] {path}")
    return path


def _write_header(ws, headers: list):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor="2D3561")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 22


def _style_sheet(ws, col_widths: list):
    from openpyxl.styles import Alignment, Border, Side
    thin = Side(style="thin", color="D0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


# ── 편의 함수 ─────────────────────────────────────────────────

def send_report(results: list[dict], date_str: str, total_collected: int,
                screened_all: list[dict] = None,
                company_name: str = "", keywords: list[str] = None,
                save: bool = True, mail: bool = True,
                subject: str = None, mail_to: str = None,
                report_prefix: str = "") -> None:
    html = build_html(results, date_str, total_collected, screened_all,
                      company_name=company_name, keywords=keywords)
    excel_path = ""
    if save:
        prefix = f"{report_prefix}_" if report_prefix else ""
        html_path  = f"report_{prefix}{date_str}.html"
        excel_path_ = f"report_{prefix}{date_str}.xlsx"
        save_html(html, date_str, path=html_path)
        if screened_all:
            excel_path = save_excel(results, screened_all, date_str, path=excel_path_)
    if mail:
        send(html, date_str, subject=subject,
             excel_path=excel_path if excel_path else None,
             mail_to=mail_to)


if __name__ == "__main__":
    import sys
    sys.stdout.reconfigure(encoding="utf-8")

    from collector import collect
    from analyzer import run_pipeline, load_profile

    date_arg = sys.argv[1] if len(sys.argv) > 1 else None
    date_str = date_arg or datetime.now().strftime("%Y%m%d")

    from config import KEYWORDS
    profile = load_profile()

    notices = collect(date_str, days_back=1)
    results, screened_all = run_pipeline(notices, profile, return_screened=True)

    subject = None

    send_report(results, date_str, total_collected=len(notices),
                screened_all=screened_all,
                company_name=profile.get("company_name", ""),
                keywords=KEYWORDS,
                save=True, mail=True,
                subject=subject)
